# -*- coding: utf-8 -*-
"""
Extractor de albaranes v2 (Seccion Salsas).
NUEVA LOGICA: en vez de mantener un maestro a mano, se cruza el Nº DE PEDIDO (45........)
que SI aparece en el albaran contra el Excel de PEDIDOS pendientes (que trae TU codigo MMPP).

Flujo:
  1) MASTER de pedidos: se acumulan las lineas del/los Excel de pedidos descargados
     (pedidos_master.csv). Cada vez que dejas un Excel nuevo de pedidos, se fusionan las
     lineas nuevas sin duplicar (clave pedido+pos+material).
  2) Se leen los PDF de albaranes, se ignora COA/lavado/cupaje/transporte, y de cada albaran
     se saca el Nº de pedido, se busca en el master y se asigna el codigo MMPP + descripcion.
  3) Se vuelca a Excel una linea por albaran, con lote/cantidad (lo mejor que da el OCR) y
     enlace al PDF archivado. Lo que no cuadra, en rojo/ambar, para revisar.
"""
import os, re, io, csv, glob, sys, shutil, unicodedata
import fitz, pytesseract
from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===================== CONFIGURACION =====================
CARPETA_PDF        = "."
CARPETA_PROCESADOS = "ALBARANES PROCESADOS"
MASTER_CSV         = "pedidos_master.csv"          # base acumulada de pedidos (la gestiona el programa)
PATRON_PEDIDOS_XLSX = "Pedidos*.xlsx"              # Excel(s) de pedidos que tu descargas y dejas en la carpeta
SALIDA_XLSX        = "Albaranes_Salsas.xlsx"
IDIOMA_OCR, DPI    = "spa", 300
# pytesseract.pytesseract.tesseract_cmd = r"C:\Tesseract\tesseract.exe"   # si hiciera falta

def _resolver_tesseract():
    base = getattr(sys, "_MEIPASS", "")
    for c in [os.path.join(base, "Tesseract-OCR", "tesseract.exe") if base else "",
              r"C:\Program Files\Tesseract-OCR\tesseract.exe",
              r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"]:
        if c and os.path.exists(c):
            pytesseract.pytesseract.tesseract_cmd = c
            td = os.path.join(os.path.dirname(c), "tessdata")
            if os.path.isdir(td): os.environ["TESSDATA_PREFIX"] = td
            return

def _norm(s):
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).upper()

# ===================== MASTER DE PEDIDOS =====================
def parsear_excel_pedidos(ruta):
    """Lee un Excel de pedidos (cabecera con 'Doc.compr.' y 'Material') y devuelve lineas."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.worksheets[0]
    col, fila_cab = {}, None
    for r in range(1, 15):
        valores = {(_norm(ws.cell(r, c).value) if ws.cell(r, c).value else ""): c
                   for c in range(1, ws.max_column + 1)}
        if "DOC.COMPR." in valores and "MATERIAL" in valores:
            fila_cab = r
            col = {"pedido": valores.get("DOC.COMPR."),
                   "proveedor": valores.get("PROVEEDOR/CENTRO SUMINISTRADOR"),
                   "material": valores.get("MATERIAL"),
                   "texto": valores.get("TEXTO BREVE"),
                   "cantidad": valores.get("CANTIDAD"),
                   "pos": valores.get("POS.")}
            break
    if not fila_cab:
        return []
    lineas = []
    for r in range(fila_cab + 1, ws.max_row + 1):
        ped = ws.cell(r, col["pedido"]).value
        mat = ws.cell(r, col["material"]).value if col["material"] else None
        if not ped or not mat:
            continue
        lineas.append({"pedido": str(ped).strip(), "material": str(mat).strip(),
                       "texto": str(ws.cell(r, col["texto"]).value or "").strip(),
                       "proveedor": str(ws.cell(r, col["proveedor"]).value or "").strip(),
                       "cantidad": str(ws.cell(r, col["cantidad"]).value or "").strip(),
                       "pos": str(ws.cell(r, col["pos"]).value or "").strip()})
    return lineas

def actualizar_master():
    """Fusiona en el master las lineas de los Excel de pedidos presentes, sin duplicar."""
    master = {}
    if os.path.exists(MASTER_CSV):
        with open(MASTER_CSV, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f, delimiter=";"):
                master[(row["pedido"], row["pos"], row["material"])] = row
    nuevos = 0
    for xls in glob.glob(os.path.join(CARPETA_PDF, PATRON_PEDIDOS_XLSX)):
        for ln in parsear_excel_pedidos(xls):
            k = (ln["pedido"], ln["pos"], ln["material"])
            if k not in master:
                master[k] = ln; nuevos += 1
    if master:
        with open(MASTER_CSV, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["pedido", "pos", "material", "texto", "proveedor", "cantidad"], delimiter=";")
            w.writeheader()
            for v in master.values():
                w.writerow({k: v.get(k, "") for k in w.fieldnames})
    indice = {}
    for v in master.values():
        indice.setdefault(v["pedido"], []).append(v)
    print(f"Master de pedidos: {len(master)} lineas ({nuevos} nuevas). Pedidos distintos: {len(indice)}")
    return indice

# ===================== CLASIFICACION =====================
def tipo_de_pagina(texto):
    t = _norm(texto)
    es_coa = (re.search(r"CERTIFICAD\w*\s+DE\s+ANALI", t) or re.search(r"CERTIFICATE\s+OF\s+ANALY", t)
              or "ANALYSIS CERTIFICATE" in t or "INFORME ANALITICO" in t)
    if es_coa:
        return "COA"
    if "EFTCO" in t or "DOCUMENTO DE LAVADO" in t or "LAVADO DE CISTERNA" in t:
        return "LAVADO"
    if "CUPAJE" in t:
        return "CUPAJE"
    if "ALBARAN TRANSPORTE" in t or "ALBARAN DE TRANSPORTE" in t:
        return "TRANSPORTE"
    if "ALBARAN" in t or "DELIVERY NOTE" in t:
        return "ALBARAN"
    return "OTRO"

# ===================== EXTRACCION =====================
def extraer_pedido(texto):
    m = re.search(r"\b(45\d{8})\b", texto)
    if m: return m.group(1)
    m = re.search(r"4\s?5\s?\d(?:\s?\d){7}", texto)
    return re.sub(r"\s", "", m.group(0)) if m else ""

def extraer_lote(texto):
    for pat in [r"N[ºo°]?\s*lote[:\s]*([A-Z0-9][A-Z0-9\-/]{3,})",
                r"LOTE/SERIE\s+(\d{1,4}/\d{2})",
                r"BATCH[:\s]*([A-Z0-9][A-Z0-9\-/]{3,})",
                r"LOTE\s+(?:BRENNTAG|FABRICANTE)\s+([A-Z0-9]{5,})",
                r"LOTE[:/\s]*([A-Z0-9][A-Z0-9\-/]{3,})"]:
        for m in re.finditer(pat, texto, re.I):
            val = m.group(1).strip(" .|")
            if re.search(r"\d", val):          # un lote real lleva algun digito
                return val
    return ""

def extraer_cantidad(texto):
    for pat in [r"(\d{1,3}(?:\.\d{3})*,\d{2})\s*KG", r"(\d{1,3}(?:\.\d{3})+)\s*KG", r"(\d+,\d{2})\s*KG"]:
        m = re.search(pat, texto, re.I)
        if m: return m.group(1)
    return ""

# ===================== EMPAREJAMIENTO =====================
def emparejar(pedido, texto_albaran, indice):
    lineas = indice.get(pedido)
    if not lineas:
        return None
    if len({l["material"] for l in lineas}) == 1:
        l = lineas[0]
        return {"material": l["material"], "texto": l["texto"], "proveedor": l["proveedor"],
                "cant_pedido": l["cantidad"], "confianza": "alta"}
    T = _norm(texto_albaran)
    def score(l):
        toks = [w for w in re.split(r"[^A-Z0-9]+", _norm(l["texto"])) if len(w) >= 4]
        return sum(1 for w in toks if w in T)
    mejor = max(lineas, key=score)
    return {"material": mejor["material"], "texto": mejor["texto"], "proveedor": mejor["proveedor"],
            "cant_pedido": mejor["cantidad"], "confianza": "media" if score(mejor) > 0 else "baja"}

# ===================== OCR =====================
def ocr_pagina(page):
    pix = page.get_pixmap(dpi=DPI)
    return pytesseract.image_to_string(Image.open(io.BytesIO(pix.tobytes("png"))), lang=IDIOMA_OCR)

def ocr_cabecera(page):
    pix = page.get_pixmap(dpi=400)
    img = Image.open(io.BytesIO(pix.tobytes("png"))); w, h = img.size
    crop = img.crop((0, 0, w, int(h * 0.30)))
    return pytesseract.image_to_string(crop, lang=IDIOMA_OCR, config="--psm 11")

# ===================== PROCESO =====================
def destino_unico(carpeta, nombre):
    base, ext = os.path.splitext(nombre)
    dest = os.path.join(carpeta, nombre); n = 2
    while os.path.exists(dest):
        dest = os.path.join(carpeta, f"{base} ({n}){ext}"); n += 1
    return dest

def procesar():
    indice = actualizar_master()
    carpeta_proc = os.path.join(CARPETA_PDF, CARPETA_PROCESADOS)
    os.makedirs(carpeta_proc, exist_ok=True)
    filas = []
    for ruta_pdf in sorted(glob.glob(os.path.join(CARPETA_PDF, "*.pdf"))):
        doc = fitz.open(ruta_pdf); filas_pdf = []
        for i, page in enumerate(doc):
            texto = ocr_pagina(page)
            if tipo_de_pagina(texto) != "ALBARAN":
                continue
            pedido = extraer_pedido(texto) or extraer_pedido(ocr_cabecera(page))
            match = emparejar(pedido, texto, indice) if pedido else None
            filas_pdf.append({"pagina": i + 1, "pedido": pedido,
                              "mmpp": (match or {}).get("material", ""),
                              "material": (match or {}).get("texto", ""),
                              "proveedor": (match or {}).get("proveedor", ""),
                              "cant_pedido": (match or {}).get("cant_pedido", ""),
                              "lote": extraer_lote(texto), "cantidad": extraer_cantidad(texto),
                              "estado": "OK" if match and match["confianza"] == "alta"
                                        else "REVISAR" if match
                                        else ("PEDIDO NO ENCONTRADO" if pedido else "SIN PEDIDO")})
        doc.close()
        if not filas_pdf:
            print(f"[AVISO] Sin albaranes en {os.path.basename(ruta_pdf)} -> se deja sin mover")
            continue
        destino = destino_unico(carpeta_proc, os.path.basename(ruta_pdf))
        rel = os.path.join(CARPETA_PROCESADOS, os.path.basename(destino))
        for f in filas_pdf:
            f.update({"archivo": os.path.basename(destino), "ruta_rel": rel}); filas.append(f)
        shutil.move(ruta_pdf, destino)
    escribir_excel(filas)
    return filas

# ===================== EXCEL =====================
def escribir_excel(filas):
    wb = Workbook(); ws = wb.active; ws.title = "Albaranes"
    cab = ["Código MMPP", "Material (pedido)", "Proveedor", "Nº Pedido", "Lote",
           "Cantidad albarán", "Cant. pedido", "Nº Albarán", "Archivo", "Pág.", "Enlace", "Estado"]
    AZUL = PatternFill("solid", start_color="1F4E78"); ROJO = PatternFill("solid", start_color="FFC7CE")
    AMBAR = PatternFill("solid", start_color="FFEB9C")
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    ws.append(cab)
    for c in range(1, len(cab) + 1):
        x = ws.cell(1, c); x.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        x.fill = AZUL; x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); x.border = borde
    ws.row_dimensions[1].height = 30
    for f in filas:
        ws.append([f["mmpp"], f["material"], f["proveedor"], f["pedido"], f["lote"],
                   f["cantidad"], f["cant_pedido"], "", f["archivo"], f["pagina"],
                   f'Ver albarán (pág. {f["pagina"]})', f["estado"]])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = Font(name="Arial", size=10); ws.cell(r, c).border = borde
        if f["estado"] != "OK":
            grave = ("NO" in f["estado"]) or (f["estado"] == "SIN PEDIDO")
            ws.cell(r, len(cab)).fill = ROJO if grave else AMBAR
            ws.cell(r, 1).fill = ROJO if not f["mmpp"] else AMBAR
        e = ws.cell(r, 11); e.hyperlink = f["ruta_rel"]
        e.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    for i, w in enumerate([12, 34, 26, 13, 16, 14, 11, 14, 24, 6, 18, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(SALIDA_XLSX)

if __name__ == "__main__":
    try:
        _resolver_tesseract()
        res = procesar()
        print(f"\nProceso terminado. Albaranes detectados: {len(res)}")
        print(f"Excel: {SALIDA_XLSX}  |  PDF movidos a: {CARPETA_PROCESADOS}")
    except Exception:
        import traceback; print("\n*** ERROR ***"); traceback.print_exc()
    input("\nPulsa Enter para cerrar...")
